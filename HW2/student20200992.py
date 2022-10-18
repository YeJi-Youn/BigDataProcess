#!/usr/bin/python3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']
totals = []

row_id = 1
for row in ws:
	if row_id != 1:
		sum = float(ws.cell(row = row_id, column = 3).value or 0) * 0.3
		sum += float(ws.cell(row = row_id, column = 4).value or 0) * 0.35
		sum += float(ws.cell(row = row_id, column = 5).value or 0) * 0.34
		sum += float(ws.cell(row = row_id, column = 6).value or 0)

		ws.cell(row = row_id, column = 7).value = sum
		totals.append(round(sum, 2))

	row_id += 1

n = len(totals)
a = int(n * 0.3)
aa = int(a / 2)
b = int(n * 0.7) - a
bb = int(b / 2)
c = n - a - b
cc = int(c / 2)

totals.sort(reverse=True)
if aa != 0:
	an = totals[aa - 1]
else:
	an = 999
a0n = totals[a - 1]
bn = totals[a + bb - 1]
b0n = totals[a + b - 1]
cn = totals[a + b + cc - 1]
c0n = totals[a + b + c - 1]

print(a, aa, b, bb, c, cc)

row_id = 2
for row in ws:
	t = float(ws.cell(row = row_id, column = 7).value or 0)
	if t >= an:
		ws.cell(row = row_id, column = 8).value = "A+"
	elif t >= a0n:
		ws.cell(row = row_id, column = 8).value = "A0"
	elif t >= bn:
		ws.cell(row = row_id, column = 8).value = "B+"
	elif t >= b0n:
		ws.cell(row = row_id, column = 8).value = "B0"
	elif t >= cn:
		ws.cell(row = row_id, column = 8).value = "C+"
	elif t >= c0n:
		ws.cell(row = row_id, column = 8).value = "C0"
	row_id += 1

wb.save("student.xlsx")
